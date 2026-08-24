from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
import re
import unicodedata

import pandas as pd

try:
    from cryptography.fernet import Fernet, InvalidToken
except ImportError:
    Fernet = None
    InvalidToken = ValueError

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None
    Alignment = Border = Font = PatternFill = Side = None
    get_column_letter = None

from src import anura_api, clientify_api, objectives, persat_api, siscor_db
from src.simple_xlsx import build_workbook as build_simple_workbook


REPORTS_DIR = Path("data/informes_semanales")
REPORT_NAME_PATTERN = re.compile(
    r"Informe_Semanal_(?:(\d{4}-\d{2}-\d{2})_al_)?(\d{4}-\d{2}-\d{2})\.xlsx$"
)
ITINERARY_PATH = Path("data/itinerario_vendedores_calle.xlsx")
ENCRYPTED_ITINERARY_PATH = Path("data/itinerario_vendedores_calle.csv.enc")
CALLS_DAILY_TARGET = 20
CONTACT_HISTORY_WEEKS = 8
PERFORMANCE_TOLERANCE = 0.05
WEEKDAY_TO_OFFSET = {
    "LUNES": 0,
    "MARTES": 1,
    "MIERCOLES": 2,
    "JUEVES": 3,
    "VIERNES": 4,
    "SABADO": 5,
    "DOMINGO": 6,
}

DISPLAY_NAMES = {
    "ZONA 13 JAVIER MOLARO": "JAVIER",
    "JONATAN MERCAO": "JONATAN",
    "JUAN C. MANZELLI": "JUAN CRUZ",
    "MACA PROTTO": "MACARENA",
    "MICAELA GONZALEZ": "MICAELA",
    "LUCIA MORENO": "LUCIA",
}

COLORS = {
    "navy": "16324F",
    "teal": "0F766E",
    "green": "16804A",
    "blue": "2563A6",
    "amber": "D97706",
    "ink": "172033",
    "muted": "5B677A",
    "line": "D7DEE8",
    "panel": "F4F7FA",
    "note": "FFF4D6",
    "white": "FFFFFF",
    "red": "B42318",
    "red_panel": "FEE2E2",
    "green_panel": "DCFCE7",
    "blue_panel": "DBEAFE",
}


@dataclass(frozen=True)
class WeeklyReportResult:
    enabled: bool
    message: str
    path: Path | None = None


def _normalize(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return "".join(char for char in text.upper() if char.isalnum())


def _display_name(zone: object) -> str:
    clean = str(zone or "").strip().upper()
    return DISPLAY_NAMES.get(clean, clean)


def _is_telemarketing(zone: object) -> bool:
    return anura_api.is_telemarketing_zone(zone)


def _anura_aliases(zone: str) -> set[str]:
    selected: set[str] = {_normalize(zone)}
    normalized_zone = _normalize(zone)
    for key, aliases in anura_api.TELEMARKETING_ACCOUNTS.items():
        candidates = (key, *aliases)
        normalized_candidates = {_normalize(value) for value in candidates}
        if any(
            normalized_zone == candidate
            or normalized_zone in candidate
            or candidate in normalized_zone
            for candidate in normalized_candidates
            if candidate
        ):
            selected.update(normalized_candidates)
    return selected


def _calls_for_zone(calls: pd.DataFrame, zone: str) -> int:
    if calls.empty:
        return 0
    aliases = _anura_aliases(zone)
    matches = pd.Series(False, index=calls.index)
    for column in ("interno", "telemarketer"):
        if column in calls.columns:
            matches |= calls[column].map(_normalize).isin(aliases)
    return int(matches.sum())


def _clientify_aliases(zone: str) -> tuple[str, ...]:
    normalized = _normalize(zone)
    if "DAVID" in normalized:
        return ("DAVID",)
    if "NOELIA" in normalized:
        return ("NOELIA",)
    if "MICAELA" in normalized:
        return ("MICAELA",)
    if any(token in normalized for token in ("MACA", "MACARENA", "PROTTO")):
        return ("MACA", "MACARENA", "PROTTO")
    if "LUCIA" in normalized:
        return ("LUCIA",)
    return (normalized,)


def _contacts_for_zone(by_owner: list[dict[str, object]], zone: str) -> int:
    aliases = _clientify_aliases(zone)
    total = 0
    for row in by_owner:
        owner = _normalize(row.get("telemarketer"))
        if any(_normalize(alias) in owner for alias in aliases if alias):
            contacts = pd.to_numeric(row.get("clientes"), errors="coerce")
            if pd.notna(contacts):
                total += int(contacts)
    return total


def _historical_contact_targets(
    period_start: date,
    cutoff: date,
    zones: tuple[str, ...],
) -> tuple[dict[str, int | None], bool, str]:
    targets: dict[str, int | None] = {zone: None for zone in zones}
    if not zones:
        return targets, True, "No hay telemarketers seleccionados."

    current_week_start = period_start - timedelta(days=period_start.weekday())
    history_end = current_week_start - timedelta(days=1)
    history_start = history_end - timedelta(days=(CONTACT_HISTORY_WEEKS * 7) - 1)
    history_sync = clientify_api.inbox_activity(
        history_start.isoformat(),
        history_end.isoformat(),
        zones,
    )
    if not history_sync.enabled:
        return targets, False, history_sync.message

    contact_totals = {zone: 0 for zone in zones}
    historical_business_days = 0.0
    for week_index in range(CONTACT_HISTORY_WEEKS):
        week_start = history_start + timedelta(days=week_index * 7)
        week_end = week_start + timedelta(days=6)
        week_result = clientify_api.inbox_activity(
            week_start.isoformat(),
            week_end.isoformat(),
            zones,
            refresh_cache=False,
        )
        if not week_result.enabled:
            return targets, False, week_result.message
        historical_business_days += objectives.business_days_between(week_start, week_end)
        for zone in zones:
            contact_totals[zone] += _contacts_for_zone(week_result.by_owner, zone)

    selected_business_days = objectives.business_days_between(period_start, cutoff)
    if not historical_business_days or not selected_business_days:
        return targets, False, "No hay días hábiles suficientes para calcular la referencia histórica."
    for zone in zones:
        daily_average = contact_totals[zone] / historical_business_days
        targets[zone] = int(round(daily_average * selected_business_days))
    return targets, True, "Promedio de las 8 semanas completas anteriores."


def _source_status(name: str, enabled: bool, message: str) -> str:
    if enabled:
        return f"{name}: OK"
    return f"{name}: no disponible ({message})"


def _performance_status(value: object) -> str:
    numeric = pd.to_numeric(value, errors="coerce")
    if pd.isna(numeric):
        return "NO EVALUABLE"
    if float(numeric) >= 1 + PERFORMANCE_TOLERANCE:
        return "POR ENCIMA"
    if float(numeric) >= 1 - PERFORMANCE_TOLERANCE:
        return "EN LINEA"
    return "POR DEBAJO"


def _read_itinerary(path: Path = ITINERARY_PATH) -> pd.DataFrame:
    columns = ["id_cliente", "vendedor", "dia_norm", "fecha_programada"]
    raw = None
    if ENCRYPTED_ITINERARY_PATH.exists() and Fernet is not None:
        key = siscor_db._snapshot_key()
        if key:
            try:
                payload = Fernet(key.encode("utf-8")).decrypt(ENCRYPTED_ITINERARY_PATH.read_bytes())
                raw = pd.read_csv(BytesIO(payload))
            except (InvalidToken, ValueError):
                raw = None
    if raw is None and path.exists():
        try:
            raw = pd.read_excel(path, sheet_name=0)
        except Exception:
            raw = None
    if raw is None:
        return pd.DataFrame(columns=columns)
    frame = raw.rename(
        columns={"Nro": "id_cliente", "Zona": "vendedor", "Dia de Visita": "dia_visita"}
    )
    required = ("id_cliente", "vendedor", "dia_visita")
    if any(column not in frame.columns for column in required):
        return pd.DataFrame(columns=columns)
    frame = frame.loc[:, required].copy()
    frame["id_cliente"] = (
        frame["id_cliente"].astype(str).str.replace(r"\.0$", "", regex=True).str.strip()
    )
    frame["vendedor"] = frame["vendedor"].fillna("").astype(str).str.strip().str.upper()
    frame["dia_norm"] = frame["dia_visita"].map(_normalize)
    frame = frame[
        frame["id_cliente"].ne("")
        & frame["vendedor"].ne("")
        & frame["dia_norm"].isin(WEEKDAY_TO_OFFSET)
    ].copy()
    return frame.drop_duplicates(["id_cliente", "vendedor", "dia_norm"])


def _planned_visits(
    itinerary: pd.DataFrame,
    zone: str,
    period_start: date,
    cutoff: date,
) -> pd.DataFrame:
    if itinerary.empty:
        return pd.DataFrame(columns=["id_cliente", "fecha_programada"])
    plan = itinerary[itinerary["vendedor"].map(_normalize).eq(_normalize(zone))].copy()
    if plan.empty:
        return pd.DataFrame(columns=["id_cliente", "fecha_programada"])
    dates_by_weekday: dict[str, list[date]] = {day: [] for day in WEEKDAY_TO_OFFSET}
    current = period_start
    while current <= cutoff:
        for day, weekday in WEEKDAY_TO_OFFSET.items():
            if current.weekday() == weekday:
                dates_by_weekday[day].append(current)
                break
        current += timedelta(days=1)

    rows = [
        {"id_cliente": item.id_cliente, "fecha_programada": scheduled}
        for item in plan.itertuples(index=False)
        for scheduled in dates_by_weekday.get(item.dia_norm, [])
    ]
    return pd.DataFrame(rows, columns=["id_cliente", "fecha_programada"]).drop_duplicates()


def _visit_performance(visits: pd.DataFrame, plan: pd.DataFrame) -> dict[str, object]:
    expected = int(len(plan)) if not plan.empty else None
    if visits.empty:
        return {
            "visitas": 0,
            "visitas_programadas": expected,
            "visitas_cumplidas": 0 if expected is not None else None,
            "visitas_fuera_itinerario": 0 if expected is not None else None,
            "visitas_sin_clasificar": 0 if expected is not None else None,
        }
    actual = visits.copy()
    actual["id_cliente"] = actual["id_cliente"].fillna("").astype(str).str.strip()
    actual["fecha_visita"] = pd.to_datetime(actual["fecha"], errors="coerce").dt.date
    actual_keys = actual.loc[
        actual["id_cliente"].ne("") & actual["fecha_visita"].notna(),
        ["id_cliente", "fecha_visita"],
    ].drop_duplicates()
    if expected is None:
        completed = outside = unclassified = None
    else:
        plan_keys = plan.rename(columns={"fecha_programada": "fecha_visita"})
        merged = actual_keys.merge(plan_keys.assign(programada=True), on=["id_cliente", "fecha_visita"], how="left")
        completed = int(merged["programada"].fillna(False).sum())
        outside = int(merged["programada"].isna().sum())
        unclassified = max(int(len(actual)) - completed - outside, 0)
    return {
        "visitas": int(len(actual)),
        "visitas_programadas": expected,
        "visitas_cumplidas": completed,
        "visitas_fuera_itinerario": outside,
        "visitas_sin_clasificar": unclassified,
    }


def _activity_data(period_start: date, cutoff: date, zones: tuple[str, ...]) -> dict[str, dict[str, object]]:
    output: dict[str, dict[str, object]] = {}
    telemarketing_zones = tuple(zone for zone in zones if _is_telemarketing(zone))
    street_zones = tuple(zone for zone in zones if not _is_telemarketing(zone))

    anura_result = anura_api.calls(period_start.isoformat(), cutoff.isoformat(), telemarketing_zones)
    contact_targets, contact_history_enabled, contact_history_message = _historical_contact_targets(
        period_start,
        cutoff,
        telemarketing_zones,
    )
    clientify_result = clientify_api.inbox_activity(
        period_start.isoformat(),
        cutoff.isoformat(),
        telemarketing_zones,
        refresh_cache=not contact_history_enabled,
    )
    persat_zones = tuple(zone for zone in street_zones if str(zone).strip().upper() in persat_api.ZONE_DEVICE_MAP)
    persat_result = persat_api.activity(period_start.isoformat(), cutoff.isoformat(), persat_zones)
    itinerary = _read_itinerary()
    expected_calls = int(objectives.business_days_between(period_start, cutoff) * CALLS_DAILY_TARGET)

    for zone in telemarketing_zones:
        calls = _calls_for_zone(anura_result.calls, zone) if anura_result.enabled else None
        calls_pace = calls / expected_calls if calls is not None and expected_calls else None
        contacts = _contacts_for_zone(clientify_result.by_owner, zone) if clientify_result.enabled else None
        expected_contacts = contact_targets.get(zone) if contact_history_enabled else None
        contacts_pace = (
            contacts / expected_contacts
            if contacts is not None and expected_contacts
            else None
        )
        output[zone] = {
            "modalidad": "Telemarketing",
            "llamadas": calls,
            "llamadas_esperadas": expected_calls,
            "ritmo_llamadas": calls_pace,
            "estado_llamadas": _performance_status(calls_pace),
            "contactos": contacts,
            "contactos_esperados": expected_contacts,
            "ritmo_contactos": contacts_pace,
            "estado_contactos": _performance_status(contacts_pace),
            "visitas": None,
            "visitas_programadas": None,
            "visitas_cumplidas": None,
            "visitas_fuera_itinerario": None,
            "visitas_sin_clasificar": None,
            "ritmo_visitas": None,
            "estado_visitas": "NO EVALUABLE",
            "ritmo_visitas_totales": None,
            "estado_visitas_totales": "NO EVALUABLE",
            "porcentaje_fuera_itinerario": None,
            "fuentes_actividad": " | ".join(
                [
                    _source_status("Anura", anura_result.enabled, anura_result.message),
                    _source_status("Clientify", clientify_result.enabled, clientify_result.message),
                    _source_status(
                        "Referencia Clientify 8 semanas",
                        contact_history_enabled,
                        contact_history_message,
                    ),
                ]
            ),
        }

    for zone in street_zones:
        device_ids = persat_api.ZONE_DEVICE_MAP.get(str(zone).strip().upper(), ())
        if not device_ids:
            output[zone] = {
                "modalidad": "Vendedor de calle",
                "llamadas": None,
                "llamadas_esperadas": None,
                "ritmo_llamadas": None,
                "estado_llamadas": "NO EVALUABLE",
                "contactos": None,
                "contactos_esperados": None,
                "ritmo_contactos": None,
                "estado_contactos": "NO EVALUABLE",
                "visitas": None,
                "visitas_programadas": None,
                "visitas_cumplidas": None,
                "visitas_fuera_itinerario": None,
                "visitas_sin_clasificar": None,
                "ritmo_visitas": None,
                "estado_visitas": "NO EVALUABLE",
                "ritmo_visitas_totales": None,
                "estado_visitas_totales": "NO EVALUABLE",
                "porcentaje_fuera_itinerario": None,
                "fuentes_actividad": "Persat: sin dispositivo asociado a esta zona",
            }
            continue
        visits = persat_result.visits
        plan = _planned_visits(itinerary, zone, period_start, cutoff)
        visit_values = {
            "visitas": None,
            "visitas_programadas": int(len(plan)) if not plan.empty else None,
            "visitas_cumplidas": None,
            "visitas_fuera_itinerario": None,
            "visitas_sin_clasificar": None,
        }
        if persat_result.enabled:
            zone_visits = visits[visits["device_id"].isin(device_ids)].copy() if not visits.empty else visits
            visit_values = _visit_performance(zone_visits, plan)
        expected_visits = visit_values["visitas_programadas"]
        completed_visits = visit_values["visitas_cumplidas"]
        total_visits = visit_values["visitas"]
        outside_visits = visit_values["visitas_fuera_itinerario"]
        visit_pace = (
            completed_visits / expected_visits
            if completed_visits is not None and expected_visits
            else None
        )
        total_visit_pace = (
            total_visits / expected_visits
            if total_visits is not None and expected_visits
            else None
        )
        outside_share = (
            outside_visits / total_visits
            if outside_visits is not None and total_visits
            else None
        )
        output[zone] = {
            "modalidad": "Vendedor de calle",
            "llamadas": None,
            "llamadas_esperadas": None,
            "ritmo_llamadas": None,
            "estado_llamadas": "NO EVALUABLE",
            "contactos": None,
            "contactos_esperados": None,
            "ritmo_contactos": None,
            "estado_contactos": "NO EVALUABLE",
            **visit_values,
            "ritmo_visitas": visit_pace,
            "estado_visitas": _performance_status(visit_pace),
            "ritmo_visitas_totales": total_visit_pace,
            "estado_visitas_totales": _performance_status(total_visit_pace),
            "porcentaje_fuera_itinerario": outside_share,
            "fuentes_actividad": (
                f"{_source_status('Persat', persat_result.enabled, persat_result.message)}"
                f" | Itinerario: {'OK' if expected_visits is not None else 'sin clientes asignados'}"
            ),
        }

    return output


def collect_report_data(cutoff: date, period_start: date | None = None) -> tuple[pd.DataFrame, date, date]:
    month_start = cutoff.replace(day=1)
    period_start = period_start or cutoff - timedelta(days=cutoff.weekday())
    if period_start > cutoff:
        raise ValueError("La fecha desde no puede ser posterior a la fecha hasta.")
    month = objectives.month_key(cutoff)
    objective_rows = objectives.load_objectives()
    objective_rows = objective_rows[objective_rows["mes"].eq(month)].copy()
    if objective_rows.empty:
        raise RuntimeError(f"No hay objetivos cargados para {month}.")

    zones = tuple(objective_rows["zona"].dropna().astype(str))
    sales = siscor_db.ventas_por_zona(month_start.isoformat(), cutoff.isoformat(), zones)
    units = siscor_db.unidades_por_zona(month_start.isoformat(), cutoff.isoformat(), zones)
    activity = _activity_data(period_start, cutoff, zones)

    base = objective_rows.loc[:, ["zona", "objetivo"]].copy()
    sales_columns = ["zona", "total", "comprobantes", "clientes"]
    sales = sales.reindex(columns=sales_columns)
    base = base.merge(sales, on="zona", how="left")
    base = base.merge(units.reindex(columns=["zona", "unidades"]), on="zona", how="left")
    for column in ("objetivo", "total", "comprobantes", "clientes", "unidades"):
        base[column] = pd.to_numeric(base[column], errors="coerce").fillna(0.0)
    base["cumplimiento"] = base.apply(
        lambda row: float(row["total"] / row["objetivo"]) if row["objetivo"] else 0.0,
        axis=1,
    )
    base["ticket_promedio"] = base.apply(
        lambda row: float(row["total"] / row["comprobantes"]) if row["comprobantes"] else 0.0,
        axis=1,
    )
    base["nombre"] = base["zona"].map(_display_name)
    month_progress = objectives.month_progress(cutoff)
    base["avance_esperado_mes"] = month_progress
    base["facturacion_esperada"] = base["objetivo"] * month_progress
    base["ritmo_ventas"] = base.apply(
        lambda row: float(row["total"] / row["facturacion_esperada"])
        if row["facturacion_esperada"]
        else None,
        axis=1,
    )
    base["estado_ventas"] = base["ritmo_ventas"].map(_performance_status)
    base["modalidad"] = base["zona"].map(lambda zone: activity[str(zone)]["modalidad"])
    for column in (
        "llamadas",
        "llamadas_esperadas",
        "ritmo_llamadas",
        "estado_llamadas",
        "contactos",
        "contactos_esperados",
        "ritmo_contactos",
        "estado_contactos",
        "visitas",
        "visitas_programadas",
        "visitas_cumplidas",
        "visitas_fuera_itinerario",
        "visitas_sin_clasificar",
        "ritmo_visitas",
        "estado_visitas",
        "ritmo_visitas_totales",
        "estado_visitas_totales",
        "porcentaje_fuera_itinerario",
        "fuentes_actividad",
    ):
        base[column] = base["zona"].map(lambda zone, key=column: activity[str(zone)][key])
    return base.sort_values("nombre").reset_index(drop=True), month_start, period_start


def _merge_value(sheet, address: str, value: object) -> None:
    sheet.merge_cells(address)
    sheet[address.split(":", 1)[0]] = value


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _style_merged(
    sheet,
    address: str,
    *,
    fill: str,
    color: str,
    size: int,
    bold: bool = False,
    horizontal: str = "center",
) -> None:
    thin = Side(style="thin", color=COLORS["line"])
    for row in sheet[address]:
        for cell in row:
            cell.fill = _fill(fill)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell = sheet[address.split(":", 1)[0]]
    cell.font = Font(name="Aptos", size=size, bold=bold, color=color)
    cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=True)


def _status_style(status: str) -> tuple[str, str]:
    if status == "POR ENCIMA":
        return COLORS["green_panel"], "166534"
    if status == "EN LINEA":
        return COLORS["blue_panel"], "1D4ED8"
    if status == "POR DEBAJO":
        return COLORS["red_panel"], COLORS["red"]
    if status == "REVISAR DATOS":
        return COLORS["red_panel"], COLORS["red"]
    if status == "SIN DIFERENCIAS":
        return COLORS["green_panel"], "166534"
    if status == "DATO DE CONTROL":
        return COLORS["note"], "8A4B08"
    return COLORS["panel"], COLORS["muted"]


def _write_metric_row(
    sheet,
    row_number: int,
    label: str,
    actual: object,
    expected: object,
    performance: object,
    status: str,
    *,
    actual_format: str = "#,##0",
    expected_format: str = "#,##0",
) -> None:
    values = [
        (f"A{row_number}:B{row_number}", label, COLORS["panel"], COLORS["ink"], "left", None),
        (f"C{row_number}:D{row_number}", actual, COLORS["white"], COLORS["ink"], "center", actual_format),
        (f"E{row_number}:F{row_number}", expected, COLORS["white"], COLORS["ink"], "center", expected_format),
        (f"G{row_number}:H{row_number}", performance, COLORS["white"], COLORS["ink"], "center", "0.0%"),
    ]
    for address, value, fill, color, horizontal, number_format in values:
        display_value = "N/D" if value is None or pd.isna(value) else value
        _merge_value(sheet, address, display_value)
        _style_merged(
            sheet,
            address,
            fill=fill,
            color=color,
            size=10,
            bold=address.startswith("A"),
            horizontal=horizontal,
        )
        if number_format and isinstance(display_value, (int, float)):
            sheet[address.split(":", 1)[0]].number_format = number_format
    fill, color = _status_style(status)
    _merge_value(sheet, f"I{row_number}:J{row_number}", status)
    _style_merged(sheet, f"I{row_number}:J{row_number}", fill=fill, color=color, size=9, bold=True)


def _seller_reading(row: pd.Series) -> str:
    sales = (
        f"Ventas {row['estado_ventas'].lower()}: lleva {float(row['ritmo_ventas']):.1%} "
        "del nivel esperado al corte."
    )
    if row["modalidad"] == "Telemarketing":
        calls = "Llamadas: dato no disponible."
        if not pd.isna(row["ritmo_llamadas"]):
            calls = (
                f"Llamadas {str(row['estado_llamadas']).lower()}: "
                f"{int(row['llamadas'])} de {int(row['llamadas_esperadas'])} esperadas."
            )
        contacts = "Contactos: referencia histórica no disponible."
        if not pd.isna(row["ritmo_contactos"]):
            contacts = (
                f"Contactos {str(row['estado_contactos']).lower()}: "
                f"{int(row['contactos'])} frente a {int(row['contactos_esperados'])} esperados "
                "según las 8 semanas anteriores."
            )
        return f"{sales} {calls} {contacts}"
    if pd.isna(row["visitas_programadas"]):
        return f"{sales} Las visitas no se califican porque esta zona no tiene itinerario cargado."
    if pd.isna(row["visitas_cumplidas"]):
        return f"{sales} Persat no estuvo disponible para evaluar el cumplimiento del itinerario."
    return (
        f"{sales} Registró {int(row['visitas'])} visitas: cumplió "
        f"{int(row['visitas_cumplidas'])} de {int(row['visitas_programadas'])} programadas, "
        f"realizó {int(row['visitas_fuera_itinerario'])} fuera del itinerario y "
        f"quedaron {int(row['visitas_sin_clasificar'])} sin clasificar."
    )


def _write_seller_sheet(sheet, row: pd.Series, cutoff: date, month_start: date, period_start: date) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A5"
    sheet.print_area = "A1:J27"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.sheet_properties.outlinePr.summaryBelow = True
    for column in range(1, 11):
        sheet.column_dimensions[get_column_letter(column)].width = 14

    _merge_value(sheet, "A1:J1", "INFORME SEMANAL DE DESEMPEÑO")
    _style_merged(sheet, "A1:J1", fill=COLORS["navy"], color=COLORS["white"], size=18, bold=True)
    _merge_value(sheet, "A2:J2", f"{row['nombre']}  |  {row['modalidad']}")
    _style_merged(sheet, "A2:J2", fill=COLORS["navy"], color=COLORS["white"], size=12, bold=True)
    period = (
        f"Corte: {cutoff:%d/%m/%Y}  |  Ventas: {month_start:%d/%m/%Y} al {cutoff:%d/%m/%Y}"
        f"  |  Actividad: {period_start:%d/%m/%Y} al {cutoff:%d/%m/%Y}"
    )
    _merge_value(sheet, "A3:J3", period)
    _style_merged(sheet, "A3:J3", fill="E8EEF5", color=COLORS["ink"], size=10)
    _merge_value(sheet, "A4:J4", f"Zona SisCor: {row['zona']}")
    _style_merged(sheet, "A4:J4", fill=COLORS["note"], color="8A4B08", size=9, bold=True)

    _merge_value(sheet, "A6:J6", "RESULTADO COMERCIAL ACUMULADO DEL MES")
    _style_merged(sheet, "A6:J6", fill=COLORS["teal"], color=COLORS["white"], size=11, bold=True, horizontal="left")
    cards = [
        ("A7:B7", "A8:B9", "Objetivo mensual", float(row["objetivo"]), COLORS["navy"], '"$"#,##0'),
        ("C7:D7", "C8:D9", "Facturación acumulada", float(row["total"]), COLORS["blue"], '"$"#,##0'),
        ("E7:F7", "E8:F9", "% de cumplimiento", float(row["cumplimiento"]), COLORS["green"], "0.0%"),
        ("G7:H7", "G8:H9", "Unidades vendidas", float(row["unidades"]), COLORS["amber"], "#,##0"),
        ("I7:J7", "I8:J9", "Ticket promedio", float(row["ticket_promedio"]), "6B5B95", '"$"#,##0'),
    ]
    for label_address, value_address, label, value, color, number_format in cards:
        _merge_value(sheet, label_address, label)
        _style_merged(sheet, label_address, fill=color, color=COLORS["white"], size=9, bold=True)
        _merge_value(sheet, value_address, value)
        _style_merged(sheet, value_address, fill=COLORS["panel"], color=COLORS["ink"], size=15, bold=True)
        sheet[value_address.split(":", 1)[0]].number_format = number_format

    expected_progress = float(row["avance_esperado_mes"])
    _merge_value(sheet, "A11:J11", f"RENDIMIENTO ESPERADO AL CORTE ({expected_progress:.1%} DEL MES HÁBIL)")
    _style_merged(sheet, "A11:J11", fill=COLORS["teal"], color=COLORS["white"], size=11, bold=True, horizontal="left")
    headers = [
        ("A12:B12", "Indicador"),
        ("C12:D12", "Resultado real"),
        ("E12:F12", "Esperado al corte"),
        ("G12:H12", "Rendimiento"),
        ("I12:J12", "Estado"),
    ]
    for address, label in headers:
        _merge_value(sheet, address, label)
        _style_merged(sheet, address, fill=COLORS["navy"], color=COLORS["white"], size=9, bold=True)
    _write_metric_row(
        sheet,
        13,
        "Facturación",
        float(row["total"]),
        float(row["facturacion_esperada"]),
        row["ritmo_ventas"],
        str(row["estado_ventas"]),
        actual_format='"$"#,##0',
        expected_format='"$"#,##0',
    )

    _merge_value(sheet, "A15:J15", "ACTIVIDAD DEL PERIODO SELECCIONADO")
    _style_merged(sheet, "A15:J15", fill=COLORS["teal"], color=COLORS["white"], size=11, bold=True, horizontal="left")
    activity_headers = [
        ("A16:B16", "Indicador"),
        ("C16:D16", "Resultado real"),
        ("E16:F16", "Base de comparación"),
        ("G16:H16", "Porcentaje"),
        ("I16:J16", "Estado"),
    ]
    for address, label in activity_headers:
        _merge_value(sheet, address, label)
        _style_merged(sheet, address, fill=COLORS["navy"], color=COLORS["white"], size=9, bold=True)
    if row["modalidad"] == "Telemarketing":
        _write_metric_row(sheet, 17, "Llamadas salientes (Anura)", row["llamadas"], row["llamadas_esperadas"], row["ritmo_llamadas"], str(row["estado_llamadas"]))
        if pd.notna(row["ritmo_contactos"]):
            _write_metric_row(sheet, 18, "Contactos únicos (Clientify)", row["contactos"], row["contactos_esperados"], row["ritmo_contactos"], str(row["estado_contactos"]))
        else:
            _write_metric_row(sheet, 18, "Contactos únicos (Clientify)", row["contactos"], row["contactos_esperados"], "-", "SIN BASE HISTÓRICA")
    elif pd.isna(row["visitas_programadas"]):
        _write_metric_row(sheet, 17, "Visitas registradas (Persat)", row["visitas"], "Sin itinerario", None, "NO EVALUABLE")
    else:
        _write_metric_row(sheet, 17, "Visitas programadas cumplidas", row["visitas_cumplidas"], row["visitas_programadas"], row["ritmo_visitas"], str(row["estado_visitas"]))
        _write_metric_row(sheet, 18, "Total de visitas realizadas", row["visitas"], row["visitas_programadas"], row["ritmo_visitas_totales"], str(row["estado_visitas_totales"]))
        _write_metric_row(sheet, 19, "Visitas fuera del itinerario", row["visitas_fuera_itinerario"], row["visitas"], row["porcentaje_fuera_itinerario"], "DATO DE CONTROL")
        unclassified = pd.to_numeric(row["visitas_sin_clasificar"], errors="coerce")
        unclassified_status = (
            "REVISAR DATOS"
            if pd.notna(unclassified) and float(unclassified) > 0
            else "SIN DIFERENCIAS"
        )
        _write_metric_row(sheet, 20, "Visitas sin clasificar", row["visitas_sin_clasificar"], "-", "-", unclassified_status)

    _merge_value(sheet, "A21:J21", "LECTURA PARA EL VENDEDOR")
    _style_merged(sheet, "A21:J21", fill=COLORS["teal"], color=COLORS["white"], size=10, bold=True, horizontal="left")
    _merge_value(sheet, "A22:J22", _seller_reading(row))
    _style_merged(sheet, "A22:J22", fill=COLORS["panel"], color=COLORS["ink"], size=9, horizontal="left")
    _merge_value(sheet, "A23:J23", "Criterio: POR ENCIMA ≥ 105% | EN LÍNEA 95% a 104,9% | POR DEBAJO < 95%.")
    _style_merged(sheet, "A23:J23", fill=COLORS["white"], color=COLORS["muted"], size=8, horizontal="left")

    _merge_value(sheet, "A25:J25", "FUENTES DEL INFORME")
    _style_merged(sheet, "A25:J25", fill="E8EEF5", color=COLORS["navy"], size=9, bold=True, horizontal="left")
    sources = f"SisCor (solo lectura): facturación, comprobantes y unidades | {row['fuentes_actividad']}"
    _merge_value(sheet, "A26:J26", sources)
    _style_merged(sheet, "A26:J26", fill=COLORS["white"], color=COLORS["muted"], size=8, horizontal="left")
    _merge_value(sheet, "A27:J27", f"Generado por Bruncas Comercial el {datetime.now():%d/%m/%Y %H:%M}")
    _style_merged(sheet, "A27:J27", fill=COLORS["white"], color=COLORS["muted"], size=8, horizontal="left")

    for row_number, height in {
        1: 30, 2: 22, 3: 21, 4: 20, 6: 21, 7: 21, 8: 25, 9: 25,
        11: 21, 12: 21, 13: 25, 15: 21, 16: 21, 17: 24, 18: 24, 19: 24, 20: 24,
        21: 21, 22: 34, 23: 19, 25: 20, 26: 27, 27: 18,
    }.items():
        sheet.row_dimensions[row_number].height = height


def build_workbook(data: pd.DataFrame, cutoff: date, month_start: date, period_start: date) -> bytes:
    if Workbook is None:
        return build_simple_workbook(data, cutoff, month_start, period_start)
    workbook = Workbook()
    workbook.remove(workbook.active)
    used_titles: set[str] = set()
    for _, row in data.iterrows():
        base_title = re.sub(r"[\\/*?:\[\]]", " ", str(row["nombre"])).strip()[:31] or "Vendedor"
        title = base_title
        suffix = 2
        while title in used_titles:
            title = f"{base_title[:27]} {suffix}"
            suffix += 1
        used_titles.add(title)
        sheet = workbook.create_sheet(title)
        _write_seller_sheet(sheet, row, cutoff, month_start, period_start)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def report_filename(cutoff: date, period_start: date | None = None) -> str:
    if period_start is None:
        return f"Informe_Semanal_{cutoff:%Y-%m-%d}.xlsx"
    return f"Informe_Semanal_{period_start:%Y-%m-%d}_al_{cutoff:%Y-%m-%d}.xlsx"


def report_period(path: Path) -> tuple[date | None, date] | None:
    match = REPORT_NAME_PATTERN.match(path.name)
    if not match:
        return None
    period_start = date.fromisoformat(match.group(1)) if match.group(1) else None
    return period_start, date.fromisoformat(match.group(2))


def generate_report(
    cutoff: date | None = None,
    output_dir: Path = REPORTS_DIR,
    period_start: date | None = None,
) -> WeeklyReportResult:
    cutoff = cutoff or date.today()
    try:
        data, month_start, effective_period_start = collect_report_data(cutoff, period_start)
        content = build_workbook(data, cutoff, month_start, effective_period_start)
        output_dir.mkdir(parents=True, exist_ok=True)
        path = output_dir / report_filename(cutoff, period_start)
        temporary = path.with_suffix(".xlsx.tmp")
        temporary.write_bytes(content)
        temporary.replace(path)
        return WeeklyReportResult(True, "OK", path)
    except Exception as exc:
        return WeeklyReportResult(False, str(exc), None)


def list_reports(output_dir: Path = REPORTS_DIR) -> list[Path]:
    if not output_dir.exists():
        return []
    return sorted(
        (path for path in output_dir.glob("Informe_Semanal_*.xlsx") if REPORT_NAME_PATTERN.match(path.name)),
        key=lambda path: (path.stat().st_mtime_ns, path.name),
        reverse=True,
    )


def latest_report(output_dir: Path = REPORTS_DIR) -> Path | None:
    reports = list_reports(output_dir)
    return reports[0] if reports else None
